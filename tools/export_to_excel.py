"""
export_to_excel.py
------------------
Exports a list of lead records to a formatted .xlsx file using openpyxl.

Columns: Name | Telefon | Adresse | Website | Rating | Bewertungen | Maps Link | Website Status
"""

import logging
import os
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

COLUMNS = [
    ("Name", 35),
    ("Telefon", 20),
    ("Adresse", 45),
    ("Website", 40),
    ("Rating", 8),
    ("Maps Link", 50),
    ("Website Status", 20),
]

FIELD_KEYS = [
    "name",
    "phone",
    "address",
    "website",
    "rating",
    "maps_link",
    "website_status",
]

HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
ROW_FILL_ALT = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
HYPERLINK_FONT = Font(color="0563C1", underline="single")
HEADER_FONT = Font(bold=True)


def _safe(value) -> str:
    """Convert None to empty string; stringify everything else."""
    if value is None:
        return ""
    return str(value)


def _build_workbook(records: list[dict]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Header row
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # Data rows
    for row_idx, record in enumerate(records, start=2):
        fill = ROW_FILL_ALT if row_idx % 2 == 0 else None

        for col_idx, key in enumerate(FIELD_KEYS, start=1):
            value = record.get(key)
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=False)

            if fill:
                cell.fill = fill

            # Website URL (col 4) — clickable hyperlink or "Keine Website"
            if key == "website":
                if value:
                    url = _safe(value)
                    cell.value = url
                    cell.hyperlink = url
                    cell.font = HYPERLINK_FONT
                else:
                    cell.value = "Keine Website"

            # Maps Link (col 7) — clickable hyperlink
            elif key == "maps_link" and value:
                url = _safe(value)
                cell.value = url
                cell.hyperlink = url
                cell.font = HYPERLINK_FONT

            # Phone — force text so Excel doesn't mangle +49 numbers
            elif key == "phone":
                cell.value = _safe(value)
                cell.number_format = "@"

            # Rating — float
            elif key == "rating" and value is not None:
                try:
                    cell.value = float(value)
                    cell.number_format = "0.0"
                except (ValueError, TypeError):
                    cell.value = _safe(value)

            # Review count — integer
            elif key == "review_count" and value is not None:
                try:
                    cell.value = int(value)
                    cell.number_format = "#,##0"
                except (ValueError, TypeError):
                    cell.value = _safe(value)

            else:
                cell.value = _safe(value)

    return wb


def export(records: list[dict], output_path: str) -> str:
    """
    Build and save the Excel workbook.
    Returns the absolute path of the saved file.
    Raises a clear error if the file is open in Excel.
    """
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = _build_workbook(records)

    try:
        wb.save(str(path))
    except PermissionError:
        raise PermissionError(
            f"Kann '{path}' nicht speichern — ist die Datei noch in Excel geöffnet? "
            "Bitte schließen und erneut ausführen."
        )

    abs_path = str(path.resolve())
    logger.info("Excel gespeichert: %s (%d Zeilen)", abs_path, len(records))
    return abs_path
